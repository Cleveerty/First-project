from flask import Flask, request, send_file
from openpyxl import load_workbook
from fpdf import FPDF
import os

app = Flask(__name__)

@app.route('/')
def index():
    # Serve your HTML file using an absolute path
    html_path = os.path.join(os.path.dirname(__file__), 'web.html')
    with open(html_path, 'r', encoding='utf-8') as f:
        return f.read()

@app.route('/convert', methods=['POST'])
def convert():
    file_path = request.form.get('filePath')
    policy_number = request.form.get('policyNumber')
    if not file_path:
        return "File path not provided.", 400

    # Make file_path absolute if it's not already
    if not os.path.isabs(file_path):
        file_path = os.path.join(os.path.dirname(__file__), file_path)

    if not os.path.exists(file_path):
        return f"File not found: {file_path}", 400

    # Load Excel file
    wb = load_workbook(file_path)
    ws = wb.active

    # Create PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Add policy number at the top
    pdf.cell(200, 10, txt=f"Policy Number: {policy_number}", ln=1)

    # Skip the first row if you don't want it in the PDF
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        row_text = "  ".join([str(cell) if cell is not None else "" for cell in row])
        pdf.cell(200, 10, txt=row_text, ln=1)

    output_pdf = os.path.join(os.path.dirname(__file__), f"{policy_number}.pdf")
    pdf.output(output_pdf)

    return send_file(output_pdf, as_attachment=True, download_name=f"{policy_number}.pdf")

if __name__ == '__main__':
    app.run(debug=True)